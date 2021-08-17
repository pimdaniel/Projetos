from openpyxl import load_workbook

path = r"C:\Claro\Desenvolvimento\test.xlsx"

wb = load_workbook(path)

#Maneiras de escrever no excel
ws =wb.worksheets[0]
ws.cell(row=1, column=1, value='SQL SERVER')
 
wb.save(path)
 
  
 
 